import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# Danh sách 11 người cần hiển thị theo yêu cầu user
TARGET_PEOPLE = ['Nhiên', 'Tân', 'Quỳnh Hà', 'Phú', 'Dũng', 'Thuỳ', 'Quyền', 'Ngân An', 'Mai', 'Đức', 'Khánh']

def split_responsibilities(wb):
    """Split "Phụ trách" into 3 separate columns for each Database sheet"""
    print("=== Splitting Phụ trách into 3 columns ===")

    # Sắp xếp thứ tự sheet
    sheet_order = ['Databse AP_MN', 'Databse AP_MB', 'Databse OB_MN (A,B)', 'Databse OB_MB (A,B,C)']

    for sheet_name in sheet_order:
        if sheet_name not in wb.sheetnames:
            print(f"Warning: {sheet_name} not found")
            continue

        ws = wb[sheet_name]
        print(f"\nProcessing {sheet_name}...")

        # Tìm row tiêu đề (row chứa "NS. Phụ trách")
        title_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_col=1, max_col=3, values_only=True), 1):
            if row[1] and "Phụ trách" in str(row[1]):
                title_row = row_idx
                break

        if not title_row:
            print(f"Warning: Title row not found in {sheet_name}")
            continue

        print(f"Title row found at {title_row}")

        # Xóa tất cả dữ liệu trong cột C, D, E (từ row title_row+1 đến cuối)
        for col in ['C', 'D', 'E']:
            for cell in ws[col]:
                if cell.row > title_row:
                    cell.value = None

        # Chuyển đổi danh sách Phụ trách thành 3 cột
        # Đối với mỗi row trong data (bỏ qua header), phân tích chuỗi phụ trách
        data_rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_col=1, max_col=2, values_only=True), 1):
            if row_idx <= title_row:
                continue

            person_in_charge = row[1]  # Cột C (dữ liệu Phụ trách)
            if person_in_charge:
                data_rows.append((row_idx, person_in_charge))

        # Phân phối vào 3 cột
        for row_idx, value in data_rows:
            name = str(value).strip()
            if name == "NS. Phụ trách":
                # Header line, bỏ qua
                continue

            # Tách họ tên nếu có nhiều tên (phân tách theo dấu phẩy, dấu |, "/", "và")
            # Xử lý các trường hợp phổ biến
            names = []
            if ',' in name:
                # "TÂN, NHIÊN" -> ["TÂN", "NHIÊN"]
                parts = [p.strip() for p in name.split(',')]
                names.extend(parts)
            elif '|' in name:
                # "PHÚ | MAI" -> ["PHÚ", "MAI"]
                parts = [p.strip() for p in name.split('|')]
                names.extend(parts)
            elif '/' in name:
                # "PHÚ/MAI" -> ["PHÚ", "MAI"]
                parts = name.split('/')
                names.extend([p.strip() for p in parts])
            elif ' và ' in name.lower():
                # "Phú và Mai" -> ["Phú", "Mai"]
                parts = [p.strip() for p in name.lower().split(' và ')]
                names.extend([p.capitalize() for p in parts])
            else:
                # Trường hợp đơn: "TÂN", "MAI", vv
                names.append(name)

            # Phân phối vào 3 cột theo vòng tròn
            names_assigned = {0: [], 1: [], 2: []}
            for i, name in enumerate(names):
                names_assigned[i % 3].append(name)

            # Cập nhật từng cột
            for col, col_num in zip(['C', 'D', 'E'], [2, 3, 4]):  # Excel columns start at 1
                ws.cell(row=row_idx, column=col_num, value='/'.join(names_assigned[col_num-2]))

        print(f"Split completed for {sheet_name}")

def rebuild_personal_progress(wb):
    """Rebuild Tiến độ cá nhân table on Dashboard with all 11 people"""
    print("\n=== Rebuilding Tiến độ cá nhân table ===")

    ws = wb['DASHBOARD']

    # Tìm vị trí bắt đầu table Tiến độ cá nhân (row 16, cols 33-48)
    start_row = 16
    start_col = 33

    # Xóa tất cả data trong 11 rows
    for row_idx in range(start_row, start_row + 11):
        for col_idx in range(start_col, start_col + 16):  # 16 columns
            ws.cell(row=row_idx, column=col_idx, value=None)

    # Vòng lặp qua từng người
    for i, person_name in enumerate(TARGET_PEOPLE):
        row_idx = start_row + i

        # Điền tên người (col 33)
        ws.cell(row=row_idx, column=start_col, value=person_name)

        # Count cho từng loại: Research, Plan B, Plan A, Deal, Done
        # Search tất cả 4 Database sheets
        research_count = 0
        plan_b_count = 0
        plan_a_count = 0
        deal_count = 0
        done_count = 0

        for sheet_name in ['Databse AP_MN', 'Databse AP_MB', 'Databse OB_MN (A,B)', 'Databse OB_MB (A,B,C)']:
            if sheet_name not in wb.sheetnames:
                continue

            ws_db = wb[sheet_name]

            # Xác định các tên có thể có trong sheet (bao gồm cả dạng viết hoa/viết thường, kết hợp)
            possible_names = [person_name]

            # Thêm các biến thể khác nếu cần
            if person_name == 'Nhiên':
                possible_names.extend(['Nhiên', 'Nhiên'])
            elif person_name == 'Tân':
                possible_names.extend(['Tân', 'Tân', 'Tân'])
            elif person_name == 'Quỳnh Hà':
                possible_names.extend(['Hà', 'Quỳnh', 'HÀ'])
            elif person_name == 'Phú':
                possible_names.extend(['Phú', 'PHÚ'])
            elif person_name == 'Dũng':
                possible_names.extend(['Dũng', 'DŨNG'])
            elif person_name == 'Thuỳ':
                possible_names.extend(['Thuỳ', 'THUỲ', 'THÙY'])
            elif person_name == 'Quyền':
                possible_names.extend(['Quyền', 'QUYỀN', 'QUYÊN'])
            elif person_name == 'Ngân An':
                possible_names.extend(['An', 'Ngân', 'AN'])
            elif person_name == 'Mai':
                possible_names.extend(['Mai', 'MAI'])
            elif person_name == 'Đức':
                possible_names.extend(['Đức', 'ĐỨC'])
            elif person_name == 'Khánh':
                possible_names.extend(['Khánh', 'KHÁNH'])

            # Count cho từng status
            for row_idx_db in range(2, 2500):  # Giả định độ rộng lên đến 2500
                status_cell = ws_db.cell(row=row_idx_db, column=18)  # Cột R
                if status_cell.value:
                    status = str(status_cell.value).strip().lower()

                    # Kiểm tra tên trong cột C (Phụ trách)
                    name_cell = ws_db.cell(row=row_idx_db, column=3)
                    name_in_sheet = str(name_cell.value).strip() if name_cell.value else ''

                    # Kiểm tra xem người này có tham gia không
                    person_found = False
                    for name in possible_names:
                        if name and name.lower() in name_in_sheet.lower():
                            person_found = True
                            break

                    if person_found:
                        # Count dựa trên status
                        if status in ['research', 'research']:
                            research_count += 1
                        elif status in ['plan b', 'plan b']:
                            plan_b_count += 1
                        elif status in ['plan a', 'plan a']:
                            plan_a_count += 1
                        elif status in ['deal', 'deal']:
                            deal_count += 1
                        elif status in ['done', 'done']:
                            done_count += 1

        # Điền vào các ô tương ứng
        col_offset = 0
        ws.cell(row=row_idx, column=start_col + 1 + col_offset, value=research_count)  # Research
        col_offset += 1
        ws.cell(row=row_idx, column=start_col + 1 + col_offset, value=plan_b_count)   # Plan B
        col_offset += 1
        ws.cell(row=row_idx, column=start_col + 1 + col_offset, value=plan_a_count)   # Plan A
        col_offset += 1
        ws.cell(row=row_idx, column=start_col + 1 + col_offset, value=deal_count)     # Deal
        col_offset += 1
        ws.cell(row=row_idx, column=start_col + 1 + col_offset, value=done_count)     # Done

        # Count tổng
        total_count = research_count + plan_b_count + plan_a_count + deal_count + done_count
        ws.cell(row=row_idx, column=start_col + 9, value=total_count)  # Target column

        # Tính toán tỷ lệ phần trăm
        if total_count > 0:
            progress_percent = (total_count / 100) * 100  # Giả định target = 100
            ws.cell(row=row_idx, column=start_col + 10, value=progress_percent)  # Tiến Độ
        else:
            ws.cell(row=row_idx, column=start_col + 10, value=0)

        print(f"  Completed {person_name}: R={research_count}, PB={plan_b_count}, PA={plan_a_count}, D={deal_count}, Dn={done_count}, Total={total_count}")

def format_dashboard(wb):
    """Format the Dashboard for better appearance"""
    print("\n=== Formatting Dashboard ===")

    ws = wb['DASHBOARD']

    # Format header for Tiến độ cá nhân
    header_row = 16
    header_start_col = 33

    # Style cho header
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

    header_labels = [
        'Người phụ trách', 'Research', 'Plan B', 'Plan A', 'Deal', 'Done',
        'Target', 'Hoàn Thành', 'Tiến Độ'
    ]

    for i, label in enumerate(header_labels):
        col = header_start_col + i
        cell = ws.cell(row=header_row, column=col)
        cell.value = label
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Format data rows
    data_start_row = 16
    data_start_col = 33

    for row_idx in range(data_start_row, data_start_row + 11):
        # Alternating row colors
        if (row_idx - data_start_row) % 2 == 0:
            for col_idx in range(data_start_col, data_start_col + 16):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        # Center align all cells in data rows
        for col_idx in range(data_start_col, data_start_col + 16):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    print("Formatting completed")

def save_workbook(wb):
    """Save the updated Excel file"""
    print("\n=== Saving workbook ===")

    # Tạo tên file với timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f'data_updated_{timestamp}.xlsx'

    wb.save(filename)
    print(f"Workbook saved as: {filename}")

    return filename

if __name__ == "__main__":
    # Load the Excel file
    wb = openpyxl.load_workbook('data.xlsx')

    # Thực hiện các tác vụ
    split_responsibilities(wb)
    rebuild_personal_progress(wb)
    format_dashboard(wb)

    # Lưu file
    filename = save_workbook(wb)

    print("\n=== All tasks completed successfully! ===")
    print(f"The updated Excel file has been saved as: {filename}")
    print("\nSummary of changes:")
    print("1. Split 'Phụ trách' into 3 columns (C, D, E) in all 4 Database sheets")
    print("2. Rebuilt 'Tiến độ cá nhân' table on Dashboard with all 11 people")
    print("3. Formatted Dashboard for better readability")
    print("4. All Research, Plan B, Plan A, Deal, and Done counts have been calculated")